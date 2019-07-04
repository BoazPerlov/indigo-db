import sys
import math
import openpyxl


dest_wb = openpyxl.load_workbook('dest.xlsx')
dest_sheet = dest_wb.active
source_wb = openpyxl.load_workbook('Source.xlsx')
source_sheet = source_wb.active

for i, row in enumerate(source_sheet.rows, start=1):
    for j, row in enumerate(dest_sheet.rows, start=1):
        if source_sheet.cell(row=i, column=2).value == dest_sheet.cell(row=j, column=1).value:
            str = source_sheet.cell(row=i, column=1).value
            dest_sheet.cell(row=j, column=8).value = str

dest_wb.save('test2.xlsx')


