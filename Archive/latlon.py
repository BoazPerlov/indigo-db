import requests
import json
import openpyxl

def getLookupPostCode(postcode):
    data = requests.get("http://api.postcodes.io/postcodes/" + postcode).text
    return data

def validatePostCode(postcode):
    postcode = postcode
    data = requests.get("http://api.postcodes.io/postcodes/" + postcode + "/validate").text
    return data

xlfile = openpyxl.load_workbook('test.xlsx')
xlfile_sheet = xlfile.active

valid_pc = 1
invalid_pc = 1
max_row= 15 #xlfile_sheet.max_row

for i in range(2,max_row+1):
    pc = xlfile_sheet.cell(row=i, column=3).value
    valid = validatePostCode(pc)
    if 'true' in valid:
        dat = json.loads(getLookupPostCode(pc))
        nested = dat.get('result')
        xlfile_sheet.cell(row=i, column=4).value = str(nested['latitude'])
        xlfile_sheet.cell(row=i, column=5).value = str(nested['longitude'])
        print("valid: " + str(valid_pc))
        valid_pc += 1
    else:
        print('invalid: ' + str(invalid_pc))
        invalid_pc += 1

xlfile.save('test.xlsx')


'''result = json.loads(getLookupPostCode('BT20 4RL'))
nested = result['result']
print(nested['latitude'])'''
